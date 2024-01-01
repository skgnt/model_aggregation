#model_aggregation.pyのサブ関数として定義
import torch
import torchvision
import tqdm
import time
import copy
import os
import glob
import random
import shutil
import glob
import pandas as pd
from PIL import Image
from matplotlib import pyplot as plt
import parameter_record as pr
import openpyxl as px
import json
import sqlite3
import sklearn.metrics as skm
import datetime
from torch.utils.tensorboard import SummaryWriter

# #データをコピーして、validationとtrainのデータを分ける関数
# def data_vt(path,result_folder="/data",split=0.8):
#     #path内のフォルダを取得
#     for folder in os.listdir(path):
#             data_path=os.path.join(path,folder)+"/*"
#             #フォルダ内のファイルを取得
#             files=glob.glob(data_path)
#             train_num=int(len(files)*split)
#             #ファイルをシャッフル
#             random.shuffle(files)
#             #フォルダのパスを作成
#             train_folder=os.path.normpath(os.path.join(result_folder,"train",folder))
#             val_folder=os.path.normpath(os.path.join(result_folder,"val",folder))
#             #フォルダを作成
#             print(os.path.abspath(train_folder))
#             os.makedirs(train_folder,exist_ok=True)
#             os.makedirs(val_folder,exist_ok=True)
#             #ファイルをコピー
#             for i in range(len(files)):
#                 if i<train_num:
#                     shutil.copy(files[i],train_folder)
#                 else:
#                     shutil.copy(files[i],val_folder)

#データをコピーして、validationとtrainとtestのデータを分ける関数
def data_vtt(path,result_folder="/data",tt_split=0.8,vtr_split=0.8):
    #path内のフォルダを取得
    for folder in os.listdir(path):
            data_path=os.path.join(path,folder)+"/**/*.*"
            #フォルダ内のファイルを取得
            files=glob.glob(data_path,recursive=True)
            test_num=int(len(files)*tt_split)
            train_num=test_num+int(((len(files)-test_num)*vtr_split))
            #ファイルをシャッフル
            random.shuffle(files)
            #フォルダのパスを作成
            train_folder=os.path.normpath(os.path.join(result_folder,"train",folder))
            val_folder=os.path.normpath(os.path.join(result_folder,"val",folder))
            test_folder=os.path.normpath(os.path.join(result_folder,"test",folder))

            print(os.path.basename(train_folder))
            #フォルダを作成
            os.makedirs(train_folder,exist_ok=True)
            os.makedirs(val_folder,exist_ok=True)
            os.makedirs(test_folder,exist_ok=True)
            #ファイルをコピー
            for i in range(len(files)):
                if i<test_num:
                    shutil.copy(files[i],test_folder)
                else:
                    if i<train_num:
                        shutil.copy(files[i],train_folder)
                    else:
                        shutil.copy(files[i],val_folder)

#モデルのトレーニングを行う関数
def train_model(model, dataloaders, loss_func, optimizer, dataset_sizes,device,run_name,
                num_epochs=25,scheduler=None,freeze=None,vt=["train","val"],save_wip_model=None,weight_only=True,log_folder="log"):
    since = time.time()

    best_model_wts = copy.deepcopy(model.state_dict())
    best_acc = 0.0

    #全てをGPUに転送
    model = model.to(device)

    epoch_size = {x: len(dataloaders[x]) for x in vt}


    os.makedirs(os.path.join("runs", run_name), exist_ok=True)
    logger = SummaryWriter(os.path.join("runs", run_name))
    # logger.add_graph(model, torch.zeros(1, 3, 224, 224).to(device))

    for epoch in range(num_epochs):
        print('Epoch {}/{}'.format(epoch, num_epochs - 1))
        print('-' * 10)

        # 最終層以外の重みを解除する
        if freeze != None:
            if epoch >= freeze:
                for param in model.parameters():
                    param.requires_grad = True
                    
                with open(f"{log_folder}/{run_name}/{run_name}_status.txt","a") as f:
                    print(f"[Epoch {epoch}/{num_epochs-1}]全ての重みを固定解除しました。")
                    print(f"[Epoch {epoch}/{num_epochs-1}]All weights have been unpinned.",file=f)
                freeze = None

        
        for switch in vt:
            if switch == 'train':
                model.train() 
            else:
                model.eval()  

            sum_loss = 0.0
            sum_correct = 0
            
            
            bar = tqdm.tqdm(dataloaders[switch])
            

            for i,(input, label) in enumerate(bar):
                
                input = input.to(device)
                label = label.to(device)

                optimizer.zero_grad()


                with torch.set_grad_enabled(switch == 'train'):
                    output = model(input)
                    _, preds = torch.max(output, 1)
                    loss = loss_func(output, label)

                    if switch == 'train':
                        loss.backward()
                        optimizer.step()

                len_batch=len(label)
                one_acc=torch.sum(preds == label.data)
                sum_loss += loss.item() * input.size(0)
                sum_correct += one_acc

                # tensorboardにlossとaccを記録
                logger.add_scalar(f'{switch}/loss', loss.item(), epoch*epoch_size[switch]+i)
                logger.add_scalar(f'{switch}/acc', one_acc.item()/len_batch, epoch*epoch_size[switch]+i)            
                
                # プログレスバーに表示
                bar.set_postfix(loss=loss.item(), acc=one_acc.item()/len_batch)

            if scheduler != None:
                if switch == 'train':
                    scheduler.step()

            epoch_loss = sum_loss / dataset_sizes[switch]
            epoch_acc = float(sum_correct) / dataset_sizes[switch]

            with open(f"{log_folder}/{run_name}/{run_name}_status.txt","a") as f:
                print(f'--{switch} Loss: {epoch_loss:.5f} Acc: {epoch_acc:.5f}')
                print(f"[Epoch {epoch}/{num_epochs-1}][{switch}] Loss: {epoch_loss:.5f} Acc: {epoch_acc:.5f}",file=f)
            
            if save_wip_model != None:
                if epoch%save_wip_model == 0:
                    os.makedirs(f"{log_folder}/{run_name}/models", exist_ok=True)
                    torch.save(model.state_dict(), f"{log_folder}/{run_name}/models/{epoch}.pth")

            # モデルをディープ・コピーします
            if switch == 'val' and epoch_acc >= best_acc:
                best_acc = epoch_acc
                best_model_wts = copy.deepcopy(model.state_dict())


    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    if len(vt)==2:
        print('Best val Acc: {:5f}'.format(best_acc))

    with open(f"{log_folder}/{run_name}/{run_name}_status.txt","a") as f:
        print('Training complete in {:.0f}m {:.0f}s'.format(
            time_elapsed // 60, time_elapsed % 60),file=f)
        if len(vt)==2:
            print('Best val Acc: {:5f}'.format(best_acc),file=f)

    
    torch.save(model.state_dict(), f"{log_folder}/{run_name}/last_model_weight.pth")
    if not weight_only:
        torch.save(model, f"{log_folder}/{run_name}/last_model_all.pth")
    

    if len(vt)==2:
        model.load_state_dict(best_model_wts)
        torch.save(model, f"{log_folder}/{run_name}/best_model_weight.pth")
        if not weight_only:
            torch.save(model.state_dict(), f"{log_folder}/{run_name}/best_model_all.pth")
    return model

#モデルの性能評価を行う関数
def test_model(model,data_path,transform,device,run_name,log_folder="log"):
    model = model.to(device)
    model.eval()
    correct = 0


    #データを読み込む
    data_imageset = ImageDatasetWithPath(root=data_path,transform=transform)
    
    idx_to_class=data_imageset.idx_to_class

    data_loader = torch.utils.data.DataLoader(data_imageset, batch_size=1, shuffle=False)

    correct = 0
    
    os.makedirs(f"{log_folder}/{run_name}",exist_ok=True)
    label_str=",".join(idx_to_class)
    #ヘッダーを書き込む
    with open(f"{log_folder}/{run_name}/{run_name}_test.csv",mode="w") as f:
        print(f"path,correct_label,{label_str}",file=f)
    bar = tqdm.tqdm(data_loader)
    with torch.no_grad():
        for data in bar:
            images, label, path = data
            images = images.to(device)
            outputs = model(images)
            #outputsに対してsoftmaxを適用
            outputs=torch.nn.functional.softmax(outputs[0],dim=0)
            
            predicted = torch.argmax(outputs)
            correct+=1 if predicted == label[0] else 0
            
            outputs=map(str,outputs.tolist())
            res=",".join(outputs)
            with open(f"{log_folder}/{run_name}/{run_name}_test.csv",mode="a") as f:
                print(f"{path[0]},{idx_to_class[label[0]]},{res}",file=f)
            
    print(f"Test Accuracy: {correct}/{len(data_loader)} ({100. * correct / len(data_loader)}%)")
    with open(f"{log_folder}/{run_name}/{run_name}_status.txt","a") as f:
        print(f"Test Accuracy: {correct}/{len(data_loader)} ({100. * correct / len(data_loader):3f}%)",file=f,end="")
            
#ImageFolderと同じ機能を持ち、かつ、画像のパスを返すようにしたDataset
class ImageDatasetWithPath(torch.utils.data.Dataset):
    def __init__(self,root,transform=None,target_transform=None):
        self.root=root
        self.transform=transform
        self.target_transform=target_transform

        #root直下のフォルダ名をクラス名とする
        classes=torchvision.datasets.ImageFolder(root=root).class_to_idx

        #直下の各フォルダにアクセスして、画像のパスを取得する。
        self.dataset=[]
        for class_name,id in classes.items():

            path=os.path.join(self.root,class_name)
            files=glob.glob(path+"/*")
            dataset_self=[(os.path.abspath(file),id) for file in files]

            self.dataset.extend(dataset_self)

        self.idx_to_class=[""]*len(classes)
        for class_name,id in classes.items():
            self.idx_to_class[id]=class_name
        
    def __len__(self):
        return len(self.dataset)
    
    def __getitem__(self,idx):
        path,label=self.dataset[idx]
        img=Image.open(path)
        if self.transform:
            img=self.transform(img)
        if self.target_transform:
            label=self.target_transform(label)

        return img,label,path

#混同行列をDataFrameで受け取り、excelに書き込む関数
def confusion_matrix(data,write_path="result.xlsx"):
    header=data.columns.values.tolist()

    openpyxl=px.Workbook()
    sheet=openpyxl.active
    sheet.title="confusion_matrix"
    sheet.merge_cells(start_row=1,start_column=1,end_row=2,end_column=2)

    sheet.cell(row=3,column=1).value="correct_label"
    sheet.cell(row=1,column=3).value="predict_label"

    #column=A~Z
    #row=1~9
    sheet.merge_cells(start_row=1,start_column=3,end_row=1,end_column=data.shape[1]+2)
    sheet.merge_cells(start_row=3,start_column=1,end_row=data.shape[0]+2,end_column=1)

    for i in range(len(header)):
        sheet.cell(row=2,column=i+3).value=header[i]
        sheet.cell(row=i+3,column=2).value=header[i]


    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            sheet.cell(row=i+3,column=j+3).value=data.iloc[i,j]
    #excelを保存
    openpyxl.save(write_path)

#2分類以上の混同行列を作成する関数
def confusion_matrix_df(data):
    
    #2列目以降のheaderを取得する。
    header=data.columns.values.tolist()[2:]
    #0でDataFrameを初期化
    df=pd.DataFrame(0,index=header,columns=header)

    for i in range(data.shape[0]):
        #2列目以降の最大値を取得して、その列のヘッダーを取得
        max_idx=data.iloc[i,2:].idxmax()
        df.loc[data.iloc[i,1],max_idx]+=1
        
    return df

#2分類の混同行列をCutOffを用いて、作成する関数
def confusion_matrix_2class(data,cutoff=0.5):
    header=data.columns.values.tolist()[-2:]
    df=pd.DataFrame(0,index=header,columns=header)

    c_p=data.loc[:,["correct_label","predict_label"]]
    for i in range(data.shape[0]):
        #2列目以降の最大値を取得して、その列のヘッダーを取得
        df.loc[c_p.iloc[i,0],c_p.iloc[i,1][:-2]]+=1


    return df
#混同行列のDataFrameを受け取り、各種指標を計算する関数
def confusion_matrix_analyze(data,sesitive_label=None,specificity_label=None):
    sensitivity=None
    specificity=None
    precision=None
    f1=None
    accuracy=None
    blanced_accuracy=None


    if sesitive_label!=None:
        #2分類の場合
        sensitivity=data.loc[sesitive_label,sesitive_label]/data.loc[sesitive_label,:].sum()
        specificity=data.loc[specificity_label,specificity_label]/data.loc[specificity_label,:].sum()
        precision=data.loc[sesitive_label,sesitive_label]/data.loc[:,sesitive_label].sum()
        f1=2*precision*sensitivity/(precision+sensitivity)
        blanced_accuracy=(sensitivity+specificity)/2
    #2分類と3分類以上のaccuracyを求める
    sum_all=0
    sum_acc=0
    for i in range(data.shape[0]):
        sum_all+=data.iloc[i,:].sum()
        sum_acc+=data.iloc[i,i]
    accuracy=sum_acc/sum_all

    analyze_dict={"sensitivity":sensitivity,"specificity":specificity,"precision":precision,"f1":f1,"accuracy":accuracy,"blanced_accuracy":blanced_accuracy}
    return analyze_dict
    
    

        


#結果の解析を行う関数
def csv_analyze(csv_path,run_name=None,setting_yaml=r"setting\plot_setting.yaml",sensitive_label=None,specificity_label=None):
    df=pd.read_csv(csv_path)

    #log_folderの最下層フォルダを取得
    log_folder=os.path.dirname(os.path.dirname(csv_path))
    
    group_name=os.path.basename(log_folder)
    group_name=group_name if group_name!="log" else "single"
    #log_folderを絶対pathに変換
    log_folder=os.path.abspath(log_folder)
    
    #2分類かどうかを判定
    if df.shape[1]<5:
        roc_data=[]
        f=3#~1,99~の小数点以下の桁数

        i=0
        fr=f+2
        s=round(1/(10**f),fr)
        c=1

        di=0.09
        flg=True

        auc=0
        #ROC曲線のデータを作成
        while True:
            # print(df[(df["correct_label"]==header[2]) & (df[header[2]] > (i/100))])
            tpr=df[(df["correct_label"]==sensitive_label) & (df[sensitive_label] > i)].shape[0]/df[df["correct_label"]==sensitive_label].shape[0]
            fpr=df[(df["correct_label"]==specificity_label) & (df[sensitive_label] > i)].shape[0]/df[df["correct_label"]==specificity_label].shape[0]
            roc_data.append([i,tpr,fpr,round(((1-fpr)+tpr)/2,5)])
            if i>=1:
                break
            if i<0.01:
                i=s*c
                c+=1
                if i==s*10:
                    s*=10
                    c=2
                i=round(i,fr)
            elif 0.01<=i<0.99:
                i+=0.01
                i=round(i,2)
            else:
                if round(di+s,fr)==s*10 and flg:
                    s=round(s/10,fr)
                    c=1

                di=s*c
                i=round(1-s*10+di,fr)
                c+=1
                # print(round(1-(1/(10**f)),fr),i)
                if round(1-(1/(10**f)),fr)==i:
                    flg=False
        
        #auc算出
        #roc_dataをpandasのDataFrameに変換
        roc_df=pd.DataFrame(roc_data,columns=["CutOff","TPR","FPR","Balanced Accuracy"])
        #roc_dfをFPFの昇順にソート
        roc_df=roc_df.sort_values(by="FPR")
        #cuttoff
        for i in range(roc_df.shape[0]-1):

           auc += (roc_df.iloc[i+1]["FPR"] - roc_df.iloc[i]["FPR"]) * roc_df.iloc[i+1]["TPR"]
        
        # print("auc",skm.auc(roc_df["FPR"],roc_df["TPR"]))
        with open(csv_path.replace(".csv","_roc.csv"),"w") as f:
            print("CutOff,TPR,FPR,Balanced Accuracy",file=f)
            for data in roc_data:
                print(f"{data[0]},{data[1]},{data[2]},{data[3]}",file=f)

        #yamlを読み込む
        ps=pr.ParameterWR()
        ps.load_yaml(setting_yaml)

        plt.clf()
        #matplotlibでROC曲線を描画
        plt.plot([i[2] for i in roc_data],[i[1] for i in roc_data])

        #グラフの範囲を指定
        plt.xlim(0,1)
        plt.ylim(0,1)

        #グラフのタイトルと軸ラベルを指定
        plt.title(ps.title)
        plt.xlabel(ps.xlabel)
        plt.ylabel(ps.ylabel)
        plt.savefig(csv_path.replace(".csv","_roc.png"))

        # plt.clf()

        # #Balanced AccuracyとCutOffの値を表示
        # plt.plot([i[0] for i in roc_data],[i[3] for i in roc_data])
        # #グラフの範囲を指定
        # plt.xlim(0,1)
        # plt.ylim(0,1)

        #Balanced Accuracyの最大値のcutoffを取得
        cutoff_ll=ps.cutoff_ll
        cutoff_ul=ps.cutoff_ul
        max_acc=0
        for i in roc_data:
            #cutOffがcutoff_ll以上,cutoff_ul以下の時のみ
            if cutoff_ll<=i[0]<=cutoff_ul:
                if max_acc<i[3]:
                    max_acc=i[3]
                    best_cutoff=i[0]
                elif max_acc==i[3]:
                    #50に近い方を選択
                    if abs(0.5-i[0])<abs(0.5-best_cutoff):
                        max_acc=i[3]
                        best_cutoff=i[0]

        print(f"Balanced Accuracyの最大値は{max_acc}で、CutOffは{best_cutoff}です。({cutoff_ll}<=CutOff<={cutoff_ul})")
        
        #dfの2列目と3列目の間にpredicted_labelを追加
        df.insert(2,"predict_label","")
        #sensitive_label>best_cutoffが正のとき,predicted_labelにsensitive_labelを代入して、そうでないときはspecificity_labelを代入する。
        df["predict_label"]=df.apply(lambda x: sensitive_label+"_p" if x[sensitive_label]>best_cutoff else specificity_label+"_p",axis=1)
        

        if run_name!=None:
            os.makedirs(f"{log_folder}/{run_name}",exist_ok=True)
            #dfをcsvに書き込む
            df.to_csv(f"{log_folder}/{run_name}/{run_name}_test_p.csv",index=False)
            with open(f"{log_folder}/{run_name}/{run_name}_status.txt","a") as f:
                print(f"The maximum value for Balanced Accuracy is {max_acc} and CutOff is {best_cutoff}.({cutoff_ll}<=CutOff<={cutoff_ul})",file=f)
        
        cutoff_cmd=confusion_matrix_2class(df,cutoff=best_cutoff)
        
        matrix_analyze=confusion_matrix_analyze(cutoff_cmd,sesitive_label=sensitive_label,specificity_label=specificity_label)
        #matrix_analyzeのkeyにcutoffを追加
        matrix_analyze["cutoff"]=best_cutoff
        #matrix_analyzeのkeyにaucを追加

        matrix_analyze["auc"]=auc
        SQLiteController(f"log/analyze_record.db").analyze_write(run_name,matrix_analyze,group=group_name,log_folder=log_folder)
        confusion_matrix(cutoff_cmd,csv_path.replace(".csv","_cm.xlsx"))
    else:
        cmd=confusion_matrix_df(df) 
        #dfの2列目と3列目の間にpredicted_labelを追加
        df.insert(2,"predict_label","")

        df["predict_label"]=df.apply(lambda x: (x[3:].idxmax())+"_p",axis=1)
        if run_name!=None:
            os.makedirs(f"{log_folder}/{run_name}",exist_ok=True)
            #dfをcsvに書き込む
            df.to_csv(f"{log_folder}/{run_name}/{run_name}_test_p.csv",index=False)
        
        matrix_analyze=confusion_matrix_analyze(cmd)       
        SQLiteController(f"log/analyze_record.db").analyze_write(run_name,matrix_analyze,group=group_name,log_folder=log_folder)
        confusion_matrix(cmd,csv_path.replace(".csv","_cm.xlsx"))

class SQLiteController:
    def __init__(self, db_path):
        """
        SQLiteControllerクラスのコンストラクターです。

        :param db_path: str, データベースファイルのパス
        """
        #db_pathにファイルが存在しない場合は、ファイルを作成する
        if not os.path.exists(db_path):
            open(db_path, "w").close()
        self.conn = sqlite3.connect(db_path)
        self.cursor = self.conn.cursor()

    def execute_query(self, query, params=None):
        """
        クエリを実行するメソッドです。

        :param query: str, 実行するクエリ
        :param params: tuple, クエリに渡すパラメーター
        """
        if params:
            self.cursor.execute(query, params)
        else:
            self.cursor.execute(query)
        self.conn.commit()

    def fetch_all(self, query, params=None):
        """
        クエリを実行し、全ての結果を取得するメソッドです。

        :param query: str, 実行するクエリ
        :param params: tuple, クエリに渡すパラメーター
        :return: list, クエリの結果
        """
        if params:
            self.cursor.execute(query, params)
        else:
            self.cursor.execute(query)
        return self.cursor.fetchall()

    def fetch_one(self, query, params=None):
        """
        クエリを実行し、1つの結果を取得するメソッドです。

        :param query: str, 実行するクエリ
        :param params: tuple, クエリに渡すパラメーター
        :return: tuple, クエリの結果
        """
        if params:
            self.cursor.execute(query, params)
        else:
            self.cursor.execute(query)
        return self.cursor.fetchone()
    
    def analyze_write(self,run_name,analyze_dict,group="single",log_folder="log"):
            """
            テスト結果をデータベースに書き込む関数。

            Parameters:
            run_name (str): テストの実行名。
            analyze_dict (dict): テスト結果の辞書。accuracy, auc, sensitivity, specificity, precision, f1, balanced_accuracy, cutoffのキーを持つ。
            group (str): テストのグループ名。デフォルトは"single"。
            log_folder (str): ログファイルの保存先。デフォルトは"log"。

            Returns:
            None
            """

            #dbにtest_resultテーブルがなければ作成
            #ただし、項目はanalize_dictのkeyに加えて、時刻(yyyy-mm-dd hh:mm:ss),run_name,group,log_folderを追加する。
            #また、idは自動で付与されるようにする。
            self.execute_query("CREATE TABLE IF NOT EXISTS test_result(id INTEGER PRIMARY KEY AUTOINCREMENT,run_name TEXT,group_name TEXT,datetime TEXT,accuracy REAL,sensitivity REAL,specificity REAL,precision REAL,f1 REAL,blanced_accuracy REAL,cutoff REAL,auc REAL,log_folder TEXT)")
            #yyyy-mm-dd hh:mm:ssの形式で時刻を取得log
            now=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            #dbに書込み
            #ただし、3分類上はanalyze_dictのaccuracy以外の値がNoneになるので、その場合はNoneは書き込まない
            if analyze_dict["blanced_accuracy"]==None:
                self.execute_query("INSERT INTO test_result(run_name,group_name,datetime,accuracy,log_folder) VALUES(?,?,?,?,?)",(run_name,group,now,analyze_dict["accuracy"],log_folder))
            else:
                self.execute_query("INSERT INTO test_result(run_name,group_name,datetime,accuracy,sensitivity,specificity,precision,f1,blanced_accuracy,cutoff,auc,log_folder) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(run_name,group,now,analyze_dict["accuracy"],analyze_dict["sensitivity"],analyze_dict["specificity"],analyze_dict["precision"],analyze_dict["f1"],analyze_dict["blanced_accuracy"],analyze_dict["cutoff"],analyze_dict["auc"],log_folder))




    def close_connection(self):
        """
        データベース接続を閉じるメソッドです。
        """
        self.conn.close()


if __name__ == "__main__":
    csv_analyze(r"log\cifar10\cifar10_efficientnet_b0\cifar10_efficientnet_b0_test.csv",run_name="cifar10_efficientnet_b0")
    #tmpのDataFrameを作成
    # df=pd.DataFrame({"a":[1,2],"b":[4,5]})
    # print(df)
    # confusion_matrix(df)